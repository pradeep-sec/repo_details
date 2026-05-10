import os
import sys
from datetime import datetime, timezone

import requests
from openpyxl import Workbook
from openpyxl.styles import Font


API_ROOT = "https://api.github.com"


def github_get(url, token, params=None):
    headers = {
        "Accept": "application/vnd.github+json",
        "Authorization": f"Bearer {token}",
        "X-GitHub-Api-Version": "2022-11-28",
        "User-Agent": "repo-export-workflow",
    }
    response = requests.get(url, headers=headers, params=params, timeout=30)
    response.raise_for_status()
    return response


def get_all_repos(token):
    repos = []
    page = 1
    while True:
        response = github_get(
            f"{API_ROOT}/user/repos",
            token,
            params={
                "per_page": 100,
                "page": page,
                "sort": "updated",
                "direction": "desc",
                "affiliation": "owner,collaborator,organization_member",
            },
        )
        data = response.json()
        if not data:
            break
        repos.extend(data)
        page += 1
    return repos


def get_last_commit_author(token, owner, repo_name, default_branch):
    if not default_branch:
        return "No default branch"

    response = github_get(
        f"{API_ROOT}/repos/{owner}/{repo_name}/commits",
        token,
        params={"per_page": 1, "sha": default_branch},
    )
    commits = response.json()
    if not commits:
        return "No commits"

    commit = commits[0]
    author_name = (
        commit.get("author", {}) or {}
    ).get("login") or commit.get("commit", {}).get("author", {}).get("name")
    return author_name or "Unknown"


def write_workbook(rows, output_path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "GitHub Repos"

    headers = [
        "Repo name",
        "repo hyper link",
        "Last committed to repo",
        "owner name of the repo",
    ]
    sheet.append(headers)

    for cell in sheet[1]:
        cell.font = Font(bold=True)

    for row_index, row in enumerate(rows, start=2):
        repo_name, repo_url, last_committer, owner_name = row
        sheet.cell(row=row_index, column=1, value=repo_name)
        link_cell = sheet.cell(row=row_index, column=2, value=repo_url)
        link_cell.hyperlink = repo_url
        link_cell.style = "Hyperlink"
        sheet.cell(row=row_index, column=3, value=last_committer)
        sheet.cell(row=row_index, column=4, value=owner_name)

    widths = {
        "A": 35,
        "B": 60,
        "C": 30,
        "D": 30,
    }
    for column, width in widths.items():
        sheet.column_dimensions[column].width = width

    workbook.save(output_path)


def main():
    token = os.environ.get("GH_ACCOUNT_READ_TOKEN")
    if not token:
        print("Missing GH_ACCOUNT_READ_TOKEN secret.", file=sys.stderr)
        sys.exit(1)

    os.makedirs("output", exist_ok=True)
    repos = get_all_repos(token)

    rows = []
    for repo in repos:
        owner_name = repo["owner"]["login"]
        repo_name = repo["name"]
        repo_url = repo["html_url"]
        default_branch = repo.get("default_branch")
        last_committer = get_last_commit_author(token, owner_name, repo_name, default_branch)
        rows.append((repo_name, repo_url, last_committer, owner_name))

    rows.sort(key=lambda item: (item[3].lower(), item[0].lower()))

    timestamp = datetime.now(timezone.utc).strftime("%Y%m%d-%H%M%S")
    output_path = os.path.join("output", f"github-repos-{timestamp}.xlsx")
    write_workbook(rows, output_path)

    print(f"Repo count: {len(rows)}")
    print(f"Excel file: {output_path}")


if __name__ == "__main__":
    main()
